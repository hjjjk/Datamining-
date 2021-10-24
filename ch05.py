import xlsxwriter
from openpyxl import *
import numpy as np
import pandas as pd
import csv
#

def process_data(path,path_1):
    df = pd.read_excel(path)
    all_df = df.iloc[:,0]
    all_df_list = all_df.tolist()
    # 获取所有数字
    all_lable =df["组织"].dropna(axis=0,how='all',inplace=False)
    all_num = [i.split(';') for i in all_df_list ]


    one_data_index = []   # 所有长度为一的值的集合
    for i in all_num:
        if len(i)==1:
            one_data_index.append(i)
    for x_res in one_data_index:
        all_num.remove(x_res)

    # 将列表中元素为一的进行改变
    all_one_data = []
    for idx in one_data_index:
        a = idx[0].split('[')[1]
        one_data = [a.split(']')[0]]
        all_one_data.append(one_data)

    all_one_data_1 = [i[0] for i in all_one_data]
    all_one_data_2 = list(set(i[0]for i in all_one_data))#单索引
    all_one_data_3 = [str([i]) for i in all_one_data_2]

    all_num_1 = [str(i) for i in all_num]        #合作数量大于一的集合
    all_num_2 = [str(i) for i in all_one_data]   #合作数量为一的集合
    all_num_x = all_num_1+all_num_2

    #去除重复值
    lable = [i.split(';') for i in set(all_df_list)]
    for it in lable:
        if len(it)==1:
            lable.remove(it)
    lable_1 = [str(i) for i in lable]
    lable_x = lable_1+all_one_data_3

    all_lable_list_1 = all_lable.tolist()  #所有标号的集合
    all_lable_list = [int(i) for i in all_lable_list_1]
    all_lable_num = len(all_lable_list)


    workbook = xlsxwriter.Workbook(path_1)
    xlsheet = workbook.add_worksheet("矩阵")
    for xp in range(all_lable_num):
        xlsheet.write(0,xp+1,all_lable_list[xp])
        xlsheet.write(xp+1,0,all_lable_list[xp])
    workbook.close()


    wb_1 = load_workbook(path_1)
    sheet_1 = wb_1["矩阵"]
    it_num = []
    for it_1 in lable_x:
        if len(it_1)<11:
            num_count = all_num_x.count(it_1)
            xp_1 = it_1.split("'")
            a = float(xp_1[1])
            it_num.append(a)
    all_one_data_4 = [float(i) for i in all_one_data_1]
    it_num_x = sorted(it_num)
    it_num_idx = [all_lable_list.index(i) for i in it_num_x]
    it_num_count = [all_one_data_4.count(i) for i in it_num_x]

    for it_1_num in range(len(it_num_x)):
        rw_1 = it_num_idx[it_1_num]+2
        count_1 = it_num_count[it_1_num]
        sheet_1.cell(rw_1,rw_1,value=count_1)
    wb_1.save(path_1)

    wb_2 = load_workbook(path_1)
    sheet_2 = wb_2["矩阵"]
    for it_2 in lable_x:
        if 13<len(it_2)<22:
            num_count_2 = all_num_x.count(it_2)
            xp_2 = it_2.split("'")
            a_2 = int(float(xp_2[1]))
            a_2_idx = all_lable_list.index(a_2)+2
            b_2 = int(float(xp_2[3]))
            b_2_idx = all_lable_list.index(a_2)+2
            sheet_2.cell(a_2_idx,b_2_idx,value=num_count_2)

    wb_2.save(path_1)

    wb_3 = load_workbook(path_1)
    sheet_3 = wb_3["矩阵"]
    for it_3 in lable_x:
        if 22<len(it_3)<33:
            num_count_3 = all_num_x.count(it_3)
            xp_3 = it_3.split("'")
            a_3 = int(float(xp_3[1]))
            a_3_idx = all_lable_list.index(a_3)+2
            b_3 = int(float(xp_3[3]))
            b_3_idx = all_lable_list.index(b_3)+2
            c_3 = int(float(xp_3[5]))
            c_3_idx = all_lable_list.index(c_3)+2
            it_3_v1 = sheet_3.cell(a_3_idx,b_3_idx).value
            it_3_v2 = sheet_3.cell(a_3_idx,c_3_idx).value
            it_3_v3 = sheet_3.cell(b_3_idx,c_3_idx).value
            if it_3_v1==None:
                count3_1 = num_count_3
            else:
                count3_1 = num_count_3 + it_3_v1
            if it_3_v2==None:
                count3_2 = num_count_3
            else:
                count3_2 = num_count_3 + it_3_v2
            if it_3_v3==None:
                count3_3 = num_count_3
            else:
                count3_3 = num_count_3 + it_3_v3
            sheet_3.cell(a_3_idx,b_3_idx,value=count3_1)
            sheet_3.cell(a_3_idx,c_3_idx,value=count3_2)
            sheet_3.cell(b_3_idx,c_3_idx,value=count3_3)
    wb_3.save(path_1)

    wb_4 = load_workbook(path_1)
    sheet_4 = wb_4["矩阵"]
    for it_4 in lable_x:
        if 33< len(it_4) <44:
            num_count_4 = all_num_x.count(it_4)
            xp_4 = it_4.split("'")
            a_4 = int(float(xp_4[1]))
            a_4_idx = all_lable_list.index(a_4)+2
            b_4 = int(float(xp_4[3]))
            b_4_idx = all_lable_list.index(b_4)+2
            c_4 = int(float(xp_4[5]))
            c_4_idx = all_lable_list.index(c_4)+2
            d_4 = int(float(xp_4[7]))
            d_4_idx = all_lable_list.index(d_4)+2
            it_4_v1 = sheet_4.cell(a_4_idx,b_4_idx).value
            it_4_v2 = sheet_4.cell(a_4_idx,c_4_idx).value
            it_4_v3 = sheet_4.cell(a_4_idx,d_4_idx).value
            it_4_v4 = sheet_4.cell(b_4_idx,c_4_idx).value
            it_4_v5 = sheet_4.cell(b_4_idx,d_4_idx).value
            it_4_v6 = sheet_4.cell(c_4_idx,d_4_idx).value
            if it_4_v1==None:
                count4_1 = num_count_4
            else:
                count4_1 = num_count_4+it_4_v1
            if it_4_v2==None:
                count4_2 = num_count_4
            else:
                count4_2 = num_count_4+it_4_v2
            if it_4_v3==None:
                count4_3 = num_count_4
            else:
                count4_3 = num_count_4+it_4_v3
            if it_4_v4==None:
                count4_4 = num_count_4
            else:
                count4_4 = num_count_4+it_4_v4
            if it_4_v5==None:
                count4_5 = num_count_4
            else:
                count4_5 = num_count_4+it_4_v5
            if it_4_v6==None:
                count4_6 = num_count_4
            else:
                count4_6 = num_count_4+it_4_v6
            sheet_4.cell(a_4_idx, b_4_idx,value=count4_1)
            sheet_4.cell(a_4_idx, c_4_idx,value=count4_2)
            sheet_4.cell(a_4_idx, d_4_idx,value=count4_3)
            sheet_4.cell(b_4_idx, c_4_idx,value=count4_4)
            sheet_4.cell(b_4_idx, d_4_idx,value=count4_5)
            sheet_4.cell(c_4_idx, d_4_idx,value=count4_6)

    wb_5 = load_workbook(path_1)
    sheet_5 = wb_5["矩阵"]
    for it_5 in lable_x:
        if 44< len(it_5) <55:
            num_count_5 = all_num_x.count(it_5)
            xp_5 = it_5.split("'")
            a_5 = int(float(xp_5[1]))
            a_5_idx = all_lable_list.index(a_5)+2
            b_5 = int(float(xp_5[3]))
            b_5_idx = all_lable_list.index(b_5)+2
            c_5 = int(float(xp_5[5]))
            c_5_idx = all_lable_list.index(c_5)+2
            d_5 = int(float(xp_5[7]))
            d_5_idx = all_lable_list.index(d_5)+2
            e_5 = int(float(xp_5[9]))
            e_5_idx = all_lable_list.index(e_5)+2
            it_5_v1 = sheet_5.cell(a_5_idx,b_5_idx).value
            it_5_v2 = sheet_5.cell(a_5_idx,c_5_idx).value
            it_5_v3 = sheet_5.cell(a_5_idx,d_5_idx).value
            it_5_v4 = sheet_5.cell(a_5_idx,e_5_idx).value
            it_5_v5 = sheet_5.cell(b_5_idx,c_5_idx).value
            it_5_v6 = sheet_5.cell(b_5_idx,d_5_idx).value
            it_5_v7 = sheet_5.cell(b_5_idx,e_5_idx).value
            it_5_v8 = sheet_5.cell(c_5_idx,d_5_idx).value
            it_5_v9 = sheet_5.cell(c_5_idx,e_5_idx).value
            it_5_v10 = sheet_5.cell(d_5_idx,e_5_idx).value

            if it_5_v1==None:
                count5_1 = num_count_5
            else:
                count5_1 = num_count_5+it_5_v1
            if it_5_v2==None:
                count5_2 = num_count_5
            else:
                count5_2 = num_count_5+it_5_v2
            if it_5_v3==None:
                count5_3 = num_count_5
            else:
                count5_3 = num_count_5+it_5_v3
            if it_5_v4==None:
                count5_4 = num_count_5
            else:
                count5_4 = num_count_5+it_5_v4
            if it_5_v5==None:
                count5_5 = num_count_5
            else:
                count5_5 = num_count_5+it_5_v5
            if it_5_v6==None:
                count5_6 = num_count_5
            else:
                count5_6 = num_count_5+it_5_v6
            if it_5_v7==None:
                count5_7 = num_count_5
            else:
                count5_7 = num_count_5+it_5_v7
            if it_5_v8==None:
                count5_8 = num_count_5
            else:
                count5_8 = num_count_5+it_5_v8
            if it_5_v9==None:
                count5_9 = num_count_5
            else:
                count5_9 = num_count_5+it_5_v9
            if it_5_v10==None:
                count5_10 = num_count_5
            else:
                count5_10 = num_count_5+it_5_v10
            sheet_5.cell(a_5_idx,b_5_idx,value=count5_1)
            sheet_5.cell(a_5_idx,c_5_idx,value=count5_2)
            sheet_5.cell(a_5_idx,d_5_idx,value=count5_3)
            sheet_5.cell(a_5_idx,e_5_idx,value=count5_4)
            sheet_5.cell(b_5_idx,c_5_idx,value=count5_5)
            sheet_5.cell(b_5_idx,d_5_idx,value=count5_6)
            sheet_5.cell(b_5_idx,e_5_idx,value=count5_7)
            sheet_5.cell(c_5_idx,d_5_idx,value=count5_8)
            sheet_5.cell(c_5_idx,e_5_idx,value=count5_9)
            sheet_5.cell(d_5_idx,e_5_idx,value=count5_10)

    wb_6 = load_workbook(path_1)
    sheet_6 = wb_6["矩阵"]
    for it_6 in lable_x:
        if 55< len(it_6) <66:
            num_count_6 = all_num_x.count(it_6)
            xp_6 = it_6.split("'")
            a_6 = int(float(xp_6[1]))
            b_6 = int(float(xp_6[3]))
            c_6 = int(float(xp_6[5]))
            d_6 = int(float(xp_6[7]))
            e_6 = int(float(xp_6[9]))
            f_6 = int(float(xp_6[11]))
            a_6_idx = all_lable_list.index(a_6)+2
            b_6_idx = all_lable_list.index(b_6)+2
            c_6_idx = all_lable_list.index(c_6)+2
            d_6_idx = all_lable_list.index(d_6)+2
            e_6_idx = all_lable_list.index(e_6)+2
            f_6_idx = all_lable_list.index(f_6)+2
            it_6_v1 = sheet_6.cell(a_6_idx, b_6_idx).value
            it_6_v2 = sheet_6.cell(a_6_idx,c_6_idx).value
            it_6_v3 = sheet_6.cell(a_6_idx,d_6_idx).value
            it_6_v4 = sheet_6.cell(a_6_idx,e_6_idx).value
            it_6_v5 = sheet_6.cell(a_6_idx,f_6_idx).value
            it_6_v6 = sheet_6.cell(b_6_idx,c_6_idx).value
            it_6_v7 = sheet_6.cell(b_6_idx,d_6_idx).value
            it_6_v8 = sheet_6.cell(b_6_idx,e_6_idx).value
            it_6_v9 = sheet_6.cell(b_6_idx,f_6_idx).value
            it_6_v10 = sheet_6.cell(c_6_idx, d_6_idx).value
            it_6_v11 = sheet_6.cell(c_6_idx, e_6_idx).value
            it_6_v12 = sheet_6.cell(c_6_idx, f_6_idx).value
            it_6_v13 = sheet_6.cell(d_6_idx, e_6_idx).value
            it_6_v14 = sheet_6.cell(d_6_idx, f_6_idx).value
            it_6_v15 = sheet_6.cell(e_6_idx, f_6_idx).value

            if it_6_v1 == None:
                count6_1 = num_count_6
            else:
                count6_1 = num_count_6+it_6_v1
            if it_6_v2 == None:
                count6_2 = num_count_6
            else:
                count6_2 = num_count_6+it_6_v2
            if it_6_v3 == None:
                count6_3= num_count_6
            else:
                count6_3= num_count_6+it_6_v3
            if it_6_v4 == None:
                count6_4= num_count_6
            else:
                count6_4= num_count_6+it_6_v4
            if it_6_v5 == None:
                count6_5= num_count_6
            else:
                count6_5= num_count_6+it_6_v5
            if it_6_v6 == None:
                count6_6= num_count_6
            else:
                count6_6= num_count_6+it_6_v6
            if it_6_v7 == None:
                count6_7= num_count_6
            else:
                count6_7= num_count_6+it_6_v7
            if it_6_v8 == None:
                count6_8= num_count_6
            else:
                count6_8= num_count_6+it_6_v8
            if it_6_v9 == None:
                count6_9= num_count_6
            else:
                count6_9= num_count_6+it_6_v9
            if it_6_v10 == None:
                count6_10= num_count_6
            else:
                count6_10= num_count_6+it_6_v10
            if it_6_v11 == None:
                count6_11= num_count_6
            else:
                count6_11= num_count_6+it_6_v11
            if it_6_v12 == None:
                count6_12= num_count_6
            else:
                count6_12= num_count_6+it_6_v12
            if it_6_v13 == None:
                count6_13= num_count_6
            else:
                count6_13= num_count_6+it_6_v13
            if it_6_v14 == None:
                count6_14= num_count_6
            else:
                count6_14= num_count_6+it_6_v14
            if it_6_v15 == None:
                count6_15= num_count_6
            else:
                count6_15= num_count_6+it_6_v15
            sheet_6.cell(a_6_idx,b_6_idx,value=count6_1)
            sheet_6.cell(a_6_idx,c_6_idx,value=count6_2)
            sheet_6.cell(a_6_idx,d_6_idx,value=count6_3)
            sheet_6.cell(a_6_idx,e_6_idx,value=count6_4)
            sheet_6.cell(a_6_idx,f_6_idx,value=count6_5)
            sheet_6.cell(b_6_idx,c_6_idx,value=count6_6)
            sheet_6.cell(b_6_idx,d_6_idx,value=count6_7)
            sheet_6.cell(b_6_idx,e_6_idx,value=count6_8)
            sheet_6.cell(b_6_idx,f_6_idx,value=count6_9)
            sheet_6.cell(c_6_idx,d_6_idx,value=count6_10)
            sheet_6.cell(c_6_idx,e_6_idx,value=count6_11)
            sheet_6.cell(c_6_idx,f_6_idx,value=count6_12)
            sheet_6.cell(d_6_idx,e_6_idx,value=count6_13)
            sheet_6.cell(d_6_idx,f_6_idx,value=count6_14)
            sheet_6.cell(e_6_idx,f_6_idx,value=it_6_v15)
    wb_6.save(path_1)


    wb_7 = load_workbook(path_1)
    sheet_7 = wb_7["矩阵"]
    for it_7 in lable_x:
        if 66 < len(it_7) < 77:
            num_count_7 = all_num_x.count(it_7)
            xp_7 = it_7.split("'")
            a_7 = int(float(xp_7[1]))
            b_7 = int(float(xp_7[3]))
            c_7 = int(float(xp_7[5]))
            d_7 = int(float(xp_7[7]))
            e_7 = int(float(xp_7[9]))
            f_7 = int(float(xp_7[11]))
            g_7 = int(float(xp_7[13]))
            a_7_idx = all_lable_list.index(a_7) + 2
            b_7_idx = all_lable_list.index(b_7) + 2
            c_7_idx = all_lable_list.index(c_7) + 2
            d_7_idx = all_lable_list.index(d_7) + 2
            e_7_idx = all_lable_list.index(e_7) + 2
            f_7_idx = all_lable_list.index(f_7) + 2
            g_7_idx = all_lable_list.index(g_7) + 2

            it_7_v1 = sheet_7.cell(a_7_idx, b_7_idx).value
            it_7_v2 = sheet_7.cell(a_7_idx, c_7_idx).value
            it_7_v3 = sheet_7.cell(a_7_idx, d_7_idx).value
            it_7_v4 = sheet_7.cell(a_7_idx, e_7_idx).value
            it_7_v5 = sheet_7.cell(a_7_idx, f_7_idx).value
            it_7_v6 = sheet_7.cell(a_7_idx, g_7_idx).value
            it_7_v7 = sheet_7.cell(b_7_idx, c_7_idx).value
            it_7_v8 = sheet_7.cell(b_7_idx, d_7_idx).value
            it_7_v9 = sheet_7.cell(b_7_idx, e_7_idx).value
            it_7_v10 = sheet_7.cell(b_7_idx, f_7_idx).value
            it_7_v11 = sheet_7.cell(b_7_idx, g_7_idx).value
            it_7_v12 = sheet_7.cell(c_7_idx, d_7_idx).value
            it_7_v13 = sheet_7.cell(c_7_idx, e_7_idx).value
            it_7_v14 = sheet_7.cell(c_7_idx, f_7_idx).value
            it_7_v15 = sheet_7.cell(c_7_idx, g_7_idx).value
            it_7_v16 = sheet_7.cell(d_7_idx, e_7_idx).value
            it_7_v17 = sheet_7.cell(d_7_idx, f_7_idx).value
            it_7_v18 = sheet_7.cell(d_7_idx, g_7_idx).value
            it_7_v19 = sheet_7.cell(e_7_idx, f_7_idx).value
            it_7_v20 = sheet_7.cell(e_7_idx, g_7_idx).value
            it_7_v21 = sheet_7.cell(f_7_idx, g_7_idx).value

            if it_7_v1 == None:
                count7_1 = num_count_7
            else:
                count7_1 = num_count_7 + it_7_v1
            if it_7_v2 == None:
                count7_2 = num_count_7
            else:
                count7_2 = num_count_7 + it_7_v2
            if it_7_v3 == None:
                count7_3 = num_count_7
            else:
                count7_3 = num_count_7 + it_7_v3
            if it_7_v4 == None:
                count7_4 = num_count_7
            else:
                count7_4 = num_count_7 + it_7_v4
            if it_7_v5 == None:
                count7_5 = num_count_7
            else:
                count7_5 = num_count_7 + it_7_v5
            if it_7_v6 == None:
                count7_6 = num_count_7
            else:
                count7_6 = num_count_7 + it_7_v6
            if it_7_v7 == None:
                count7_7 = num_count_7
            else:
                count7_7 = num_count_7 + it_7_v7
            if it_7_v8 == None:
                count7_8 = num_count_7
            else:
                count7_8 = num_count_7 + it_7_v8
            if it_7_v9 == None:
                count7_9 = num_count_7
            else:
                count7_9 = num_count_7 + it_7_v9
            if it_7_v10 == None:
                count7_10 = num_count_7
            else:
                count7_10 = num_count_7 + it_7_v10
            if it_7_v11 == None:
                count7_11 = num_count_7
            else:
                count7_11 = num_count_7 + it_7_v11
            if it_7_v12 == None:
                count7_12 = num_count_7
            else:
                count7_12 = num_count_7 + it_7_v12
            if it_7_v13 == None:
                count7_13 = num_count_7
            else:
                count7_13 = num_count_7 + it_7_v13
            if it_7_v14 == None:
                count7_14 = num_count_7
            else:
                count7_14 = num_count_7 + it_7_v14
            if it_7_v15 == None:
                count7_15 = num_count_7
            else:
                count7_15 = num_count_7 + it_7_v15
            if it_7_v16 == None:
                count7_16 = num_count_7
            else:
                count7_16 = num_count_7 + it_7_v16
            if it_7_v17 == None:
                count7_17 = num_count_7
            else:
                count7_17 = num_count_7 + it_7_v17
            if it_7_v18 == None:
                count7_18 = num_count_7
            else:
                count7_18 = num_count_7 + it_7_v18
            if it_7_v19 == None:
                count7_19 = num_count_7
            else:
                count7_19 = num_count_7 + it_7_v1
            if it_7_v20 == None:
                count7_20 = num_count_7
            else:
                count7_20 = num_count_7 + it_7_v20
            if it_7_v21 == None:
                count7_21 = num_count_7
            else:
                count7_21 = num_count_7 + it_7_v21

            sheet_7.cell(a_7_idx, b_7_idx, value=count7_1)
            sheet_7.cell(a_7_idx, c_7_idx, value=count7_2)
            sheet_7.cell(a_7_idx, d_7_idx, value=count7_3)
            sheet_7.cell(a_7_idx, e_7_idx, value=count7_4)
            sheet_7.cell(a_7_idx, f_7_idx, value=count7_5)
            sheet_7.cell(a_7_idx, g_7_idx, value=count7_6)
            sheet_7.cell(b_7_idx, c_7_idx, value=count7_7)
            sheet_7.cell(b_7_idx, d_7_idx, value=count7_8)
            sheet_7.cell(b_7_idx, e_7_idx, value=count7_9)
            sheet_7.cell(b_7_idx, f_7_idx, value=count7_10)
            sheet_7.cell(b_7_idx, g_7_idx, value=count7_11)
            sheet_7.cell(c_7_idx, d_7_idx, value=count7_12)
            sheet_7.cell(c_7_idx, e_7_idx, value=count7_13)
            sheet_7.cell(c_7_idx, f_7_idx, value=count7_14)
            sheet_7.cell(c_7_idx, g_7_idx, value=count7_15)
            sheet_7.cell(d_7_idx, e_7_idx, value=count7_16)
            sheet_7.cell(d_7_idx, f_7_idx, value=count7_17)
            sheet_7.cell(d_7_idx, g_7_idx, value=count7_18)
            sheet_7.cell(e_7_idx, f_7_idx, value=count7_19)
            sheet_7.cell(e_7_idx, g_7_idx, value=count7_20)
            sheet_7.cell(f_7_idx, g_7_idx, value=count7_21)
    wb_7.save(path_1)




path = 'D:/挖掘/实验一：数据及数据预处理/电子信息产业原始数据/1电子信息材料/2009-2014合作编码(两两组合).xls'
path_1 = r'D:/挖掘/实验一：数据及数据预处理/电子信息产业原始数据/1电子信息材料/2009-2014(矩阵).xlsx'

process_data(path,path_1)