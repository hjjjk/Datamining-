import xlsxwriter
from openpyxl import *
import numpy as np
import pandas as pd
import csv
#
path = 'D:/挖掘/实验一：数据及数据预处理/电子信息产业原始数据/1电子信息材料/2009-2014合作编码(两两组合).xls'

path_1 = r'D:/挖掘/实验一：数据及数据预处理/电子信息产业原始数据/1电子信息材料/2009-2014(矩阵).xlsx'

df = pd.read_excel(path)
all_df = df.iloc[:,0]
all_df_list = all_df.tolist()
# 获取所有数字
all_lable =df["组织"].dropna(axis=0,how='all',inplace=False)
all_num = [i.split(';') for i in all_df_list ]



one_data_index = []   # 所有长度为一的值的集合
for i in all_num:
    if len(i)==1:
        one_data_index.append(i)
for x_res in one_data_index:
    all_num.remove(x_res)

# 将列表中元素为一的进行改变
all_one_data = []
for idx in one_data_index:
    a = idx[0].split('[')[1]
    one_data = [a.split(']')[0]]
    all_one_data.append(one_data)

all_one_data_1 = [i[0] for i in all_one_data]
all_one_data_2 = list(set(i[0]for i in all_one_data))#单索引
all_one_data_3 = [str([i]) for i in all_one_data_2]

all_num_1 = [str(i) for i in all_num]        #合作数量大于一的集合
all_num_2 = [str(i) for i in all_one_data]   #合作数量为一的集合
all_num_x = all_num_1+all_num_2

#去除重复值
lable = [i.split(';') for i in set(all_df_list)]
for it in lable:
    if len(it)==1:
        lable.remove(it)
lable_1 = [str(i) for i in lable]
lable_x = lable_1+all_one_data_3

all_lable_list_1 = all_lable.tolist()  #所有标号的集合
all_lable_list = [int(i) for i in all_lable_list_1]
all_lable_num = len(all_lable_list)

wb = load_workbook(path_1)
sheet = wb["矩阵"]
for xp in range(all_lable_num):
    pp = all_lable_list[xp]
    rw =xp+2
    sheet.cell(1,rw,value=pp)
    sheet.cell(rw,1,value=pp)
wb.save(path_1)

# wb = load_workbook(path_1)
# sheet = wb["矩阵"]
